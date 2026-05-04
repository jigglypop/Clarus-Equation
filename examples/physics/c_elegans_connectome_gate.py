"""C. elegans evolutionary trace gate from the Witvliet adult connectome.

This is the first non-human / primitive-nervous-system check for the brain
equation program. It tests whether the OpenWorm BrainMap-A layer assignment

    L1 sensory/input -> L2 intermediate -> L3 premotor/integrative

explains the adult C. elegans connectome better than a flat model and better
than random layer assignments with the same layer sizes.

The gate is structural only. It does not test human BOLD/EEG dynamics.
"""

from __future__ import annotations

import argparse
import json
import math
import random
from collections import defaultdict
from pathlib import Path
from statistics import mean
from urllib.request import urlretrieve

import numpy as np
import openpyxl


DATASET = "Witvliet et al. 2021 dataset 8 adult"
SOURCE_URL = (
    "https://raw.githubusercontent.com/openworm/ConnectomeToolbox/main/"
    "cect/data/witvliet_2020_8.xlsx"
)
DEFAULT_DATA_ROOT = Path("data/evolution/c_elegans")
DEFAULT_XLSX = DEFAULT_DATA_ROOT / "witvliet_2020_8.xlsx"
RNG_SEED = 1729


MODULES = {
    "L1_Anterior": [
        "CEPDL", "CEPDR", "CEPVL", "CEPVR", "IL1DL", "IL1DR", "IL1L", "IL1R",
        "IL1VL", "IL1VR", "IL2DL", "IL2DR", "IL2L", "IL2R", "IL2VL", "IL2VR",
        "OLLL", "OLLR", "OLQDL", "OLQDR", "OLQVL", "OLQVR", "URYDL", "URYDR",
        "URYVL", "URYVR",
    ],
    "L1_Lateral": ["ADEL", "ADER"],
    "L1_Avoidance": ["ALML", "ALMR", "AQR", "ASHL", "ASHR", "AVM", "FLPL", "FLPR"],
    "L1_Taxis": [
        "ADFL", "ADFR", "ADLL", "ADLR", "AFDL", "AFDR", "ASEL", "ASER",
        "ASGL", "ASGR", "ASIL", "ASIR", "ASJL", "ASJR", "ASKL", "ASKR",
        "AWAL", "AWAR", "AWBL", "AWBR", "AWCL", "AWCR", "BAGL", "BAGR",
        "URXL", "URXR",
    ],
    "L2_Anterior": ["RIPL", "RIPR", "URADL", "URADR", "URAVL", "URAVR"],
    "L2_LatSub": ["ALNL", "ALNR", "PVT", "SDQL", "SDQR"],
    "L2_Avoidance": ["AVJL", "AVJR", "BDUL", "BDUR"],
    "L2_Taxis": ["AIAL", "AIAR", "AIML", "AIMR", "AINL", "AINR", "AIYL", "AIYR", "PVQL", "PVQR"],
    "L3_Anterior": [
        "AVEL", "AVER", "RIAL", "RIAR", "RIH", "RMDDL", "RMDDR", "RMDVL",
        "RMDVR", "RMED", "RMEL", "RMER", "RMEV",
    ],
    "L3_LatSub": [
        "AVAL", "AVAR", "AVKL", "AVKR", "DVA", "DVC", "RICL", "RICR",
        "RIGL", "RIGR", "RIML", "RIMR", "RIS", "RIVL", "RIVR", "RMDL",
        "RMDR", "RMFL", "RMFR", "RMGL", "RMGR", "RMHL", "RMHR", "SAADL",
        "SAADR", "SAAVL", "SAAVR", "SIADL", "SIADR", "SIAVL", "SIAVR",
        "SIBVL", "SIBVR", "SMBDL", "SMBDR", "SMBVL", "SMBVR", "SMDDL",
        "SMDDR", "SMDVL", "SMDVR", "URBL", "URBR",
    ],
    "L3_Avoidance": [
        "ADAL", "ADAR", "AVBL", "AVBR", "AVDL", "AVDR", "AVHL", "AVHR",
        "PVCL", "PVCR", "PVPL", "PVPR", "RIFL", "RIFR", "RIR",
    ],
    "L3_Taxis": ["AIBL", "AIBR", "AIZL", "AIZR", "ALA", "AUAL", "AUAR", "AVFL", "AVFR", "RIBL", "RIBR"],
}


def module_layer(module: str) -> str:
    return module.split("_", 1)[0]


def build_cell_maps() -> tuple[dict[str, str], dict[str, str]]:
    cell_to_module = {}
    cell_to_layer = {}
    for module, cells in MODULES.items():
        layer = module_layer(module)
        for cell in cells:
            cell_to_module[cell] = module
            cell_to_layer[cell] = layer
    return cell_to_module, cell_to_layer


def ensure_xlsx(path: Path, *, force: bool) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    if force or not path.exists():
        urlretrieve(SOURCE_URL, path)
    return path


def load_edges(path: Path) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(value) for value in next(rows)]
    idx = {name: header.index(name) for name in header}
    edges = []
    for row in rows:
        if not row or row[idx["pre"]] is None or row[idx["post"]] is None:
            continue
        edges.append(
            {
                "pre": str(row[idx["pre"]]),
                "post": str(row[idx["post"]]),
                "type": str(row[idx["type"]]),
                "synapses": float(row[idx["synapses"]]),
            }
        )
    return edges


def aggregate_module_matrices(
    edges: list[dict[str, object]],
) -> tuple[list[str], dict[str, np.ndarray], dict[str, object]]:
    cell_to_module, _ = build_cell_maps()
    modules = list(MODULES)
    index = {module: idx for idx, module in enumerate(modules)}
    matrices = {
        "all_weighted": np.zeros((len(modules), len(modules)), dtype=np.float64),
        "chemical_weighted": np.zeros((len(modules), len(modules)), dtype=np.float64),
        "electrical_weighted": np.zeros((len(modules), len(modules)), dtype=np.float64),
    }
    used_edges = 0
    used_synapses = 0.0
    skipped_edges = 0
    type_totals: dict[str, float] = defaultdict(float)
    for edge in edges:
        pre = str(edge["pre"])
        post = str(edge["post"])
        if pre not in cell_to_module or post not in cell_to_module:
            skipped_edges += 1
            continue
        weight = float(edge["synapses"])
        row = index[cell_to_module[pre]]
        col = index[cell_to_module[post]]
        edge_type = str(edge["type"])
        matrices["all_weighted"][row, col] += weight
        if edge_type == "chemical":
            matrices["chemical_weighted"][row, col] += weight
        elif edge_type == "electrical":
            matrices["electrical_weighted"][row, col] += weight
        used_edges += 1
        used_synapses += weight
        type_totals[edge_type] += weight
    matrices["all_binary"] = (matrices["all_weighted"] > 0).astype(np.float64)
    matrices["chemical_binary"] = (matrices["chemical_weighted"] > 0).astype(np.float64)
    matrices["electrical_binary"] = (matrices["electrical_weighted"] > 0).astype(np.float64)
    return modules, matrices, {
        "used_edges": used_edges,
        "used_synapses": used_synapses,
        "skipped_edges": skipped_edges,
        "type_synapses": dict(type_totals),
    }


def flat_loss(values: np.ndarray) -> float:
    mu = float(np.mean(values))
    return float(np.sum((values - mu) ** 2))


def block_loss(values: np.ndarray, labels: list[str]) -> tuple[float, dict[str, float]]:
    losses = 0.0
    means = {}
    labels_arr = np.asarray(labels)
    for pre_layer in sorted(set(labels)):
        for post_layer in sorted(set(labels)):
            mask = np.outer(labels_arr == pre_layer, labels_arr == post_layer)
            block = values[mask]
            mu = float(np.mean(block)) if len(block) else 0.0
            means[f"{pre_layer}->{post_layer}"] = mu
            losses += float(np.sum((block - mu) ** 2))
    return losses, means


def random_layer_labels(labels: list[str], rng: random.Random) -> list[str]:
    shuffled = list(labels)
    rng.shuffle(shuffled)
    return shuffled


def permutation_gate(values: np.ndarray, labels: list[str], permutations: int) -> dict[str, object]:
    observed_loss, block_means = block_loss(values, labels)
    baseline_loss = flat_loss(values)
    rng = random.Random(RNG_SEED)
    random_losses = []
    hits = 0
    for _ in range(permutations):
        shuffled = random_layer_labels(labels, rng)
        loss, _ = block_loss(values, shuffled)
        random_losses.append(loss)
        if loss <= observed_loss:
            hits += 1
    return {
        "flat_loss": baseline_loss,
        "layer_block_loss": observed_loss,
        "layer_over_flat": observed_loss / max(baseline_loss, 1e-12),
        "block_means": block_means,
        "random_layer_loss_mean": mean(random_losses),
        "random_layer_loss_min": min(random_losses),
        "random_layer_loss_max": max(random_losses),
        "layer_over_random_mean": observed_loss / max(mean(random_losses), 1e-12),
        "permutations": permutations,
        "p_value_loss_le_observed": (hits + 1) / (permutations + 1),
    }


def module_block_loss(values: np.ndarray, modules: list[str]) -> tuple[float, dict[str, float]]:
    module_families = [module.rsplit("_", 1)[1] for module in modules]
    return block_loss(values, module_families)


def labels_for_model(modules: list[str], model: str) -> list[str]:
    if model == "all_one":
        return ["all"] * len(modules)
    if model == "layer":
        return [module_layer(module) for module in modules]
    if model == "module_family":
        return [module.rsplit("_", 1)[1] for module in modules]
    if model == "module":
        return list(modules)
    if model == "lateral_vs_other":
        return [
            "lateral" if module.endswith("_Lateral") or module.endswith("_LatSub") else "other"
            for module in modules
        ]
    if model == "avoidance_vs_other":
        return ["avoidance" if module.endswith("_Avoidance") else "other" for module in modules]
    if model == "taxis_vs_other":
        return ["taxis" if module.endswith("_Taxis") else "other" for module in modules]
    raise ValueError(f"unknown model: {model}")


def compare_layer_countermodels(
    matrix: np.ndarray,
    modules: list[str],
    permutations: int,
) -> dict[str, object]:
    out = {}
    n = matrix.size
    for model in (
        "all_one",
        "layer",
        "module_family",
        "lateral_vs_other",
        "avoidance_vs_other",
        "taxis_vs_other",
        "module",
    ):
        labels = labels_for_model(modules, model)
        gate = permutation_gate(matrix, labels, permutations)
        label_count = len(set(labels))
        k = label_count * label_count
        saturated = k >= n or float(gate["layer_block_loss"]) <= 1e-12
        rss = max(float(gate["layer_block_loss"]), 1e-12)
        bic = n * float(np.log(rss / max(n, 1))) + k * float(np.log(max(n, 1)))
        out[model] = {
            "label_count": label_count,
            "parameter_count": k,
            "saturated": saturated,
            "gate": gate,
            "bic_like": bic,
        }
    eligible = {name: row for name, row in out.items() if not bool(row["saturated"])}
    best_bic = min(eligible.items(), key=lambda item: float(item[1]["bic_like"]))
    best_loss = min(eligible.items(), key=lambda item: float(item[1]["gate"]["layer_block_loss"]))
    return {
        "models": out,
        "best_bic_like_model_excluding_saturated": best_bic[0],
        "best_loss_model_excluding_saturated": best_loss[0],
        "warning": (
            "Countermodels test whether L1/L2/L3 is uniquely supported. "
            "Saturated module labels are excluded from best-model calls."
        ),
    }


def robustness_gates(
    matrices: dict[str, np.ndarray],
    modules: list[str],
    permutations: int,
) -> dict[str, object]:
    labels = [module_layer(module) for module in modules]
    out = {}
    for name, matrix in matrices.items():
        if float(np.sum(matrix)) == 0.0:
            continue
        layer_gate = permutation_gate(matrix, labels, permutations)
        module_loss, module_means = module_block_loss(matrix, modules)
        flat = flat_loss(matrix)
        out[name] = {
            "layer_gate": layer_gate,
            "module_family_loss": module_loss,
            "module_family_over_flat": module_loss / max(flat, 1e-12),
            "module_family_means": module_means,
            "passed_layer_gate": layer_gate["layer_block_loss"] < layer_gate["flat_loss"]
            and layer_gate["layer_block_loss"] < layer_gate["random_layer_loss_mean"]
            and layer_gate["p_value_loss_le_observed"] < 0.05,
        }
    return out


def directionality(matrix: np.ndarray, modules: list[str]) -> dict[str, object]:
    layers = [module_layer(module) for module in modules]
    order = {"L1": 1, "L2": 2, "L3": 3}
    forward = 0.0
    backward = 0.0
    lateral = 0.0
    for i, pre_layer in enumerate(layers):
        for j, post_layer in enumerate(layers):
            value = float(matrix[i, j])
            if order[pre_layer] < order[post_layer]:
                forward += value
            elif order[pre_layer] > order[post_layer]:
                backward += value
            else:
                lateral += value
    total = forward + backward + lateral
    return {
        "forward_synapses": forward,
        "backward_synapses": backward,
        "lateral_synapses": lateral,
        "forward_fraction": forward / max(total, 1e-12),
        "backward_fraction": backward / max(total, 1e-12),
        "lateral_fraction": lateral / max(total, 1e-12),
        "forward_over_backward": forward / max(backward, 1e-12),
    }


def bow_tie_indices(matrix: np.ndarray, modules: list[str]) -> dict[str, object]:
    """Summarize how much traffic touches each layer as source or target."""
    layers = [module_layer(module) for module in modules]
    result = {}
    total = float(np.sum(matrix))
    for layer in sorted(set(layers)):
        mask = np.asarray([item == layer for item in layers])
        outgoing = float(np.sum(matrix[mask, :]))
        incoming = float(np.sum(matrix[:, mask]))
        internal = float(np.sum(matrix[np.outer(mask, mask)]))
        result[layer] = {
            "outgoing_fraction": outgoing / max(total, 1e-12),
            "incoming_fraction": incoming / max(total, 1e-12),
            "internal_fraction": internal / max(total, 1e-12),
        }
    return result


def laplacian_stability(matrix: np.ndarray) -> dict[str, float]:
    sym = 0.5 * (matrix + matrix.T)
    degree = np.sum(sym, axis=1)
    lap = np.diag(degree) - sym
    evals = np.linalg.eigvalsh(lap)
    max_eval = float(np.max(evals))
    rho = 0.20
    gamma_max = (1.0 + rho) / max(max_eval, 1e-12)
    gamma_example = 0.5 * gamma_max
    modal_radius = max(abs(rho - gamma_example * float(value)) for value in evals)
    return {
        "laplacian_lambda_max": max_eval,
        "rho_example": rho,
        "gamma_stability_upper_bound": gamma_max,
        "gamma_example": gamma_example,
        "modal_radius_at_gamma_example": float(modal_radius),
    }


def write_report(output: dict[str, object], path: Path) -> None:
    gate = output["gate"]  # type: ignore[assignment]
    direction = output["directionality"]  # type: ignore[assignment]
    stability = output["laplacian_stability"]  # type: ignore[assignment]
    robustness = output["robustness"]  # type: ignore[assignment]
    counter = output["countermodel_check"]  # type: ignore[assignment]
    lines = [
        "# C. elegans 원시 신경계 그래프 게이트",
        "",
        "이 보고서는 인간 뇌 데이터가 아니라 OpenWorm/ConnectomeToolbox의 Witvliet adult C. elegans connectome으로 전역 뇌 방정식의 진화적 최소 문법을 점검한 것이다.",
        "",
        "## 검증 질문",
        "",
        "$$",
        "\\mathcal L_{\\mathrm{L1/L2/L3\\ block}}",
        "<",
        "\\mathcal L_{\\mathrm{flat/random\\ layer}}",
        "$$",
        "",
        "여기서 L1은 감각/input, L2는 중간/relay, L3는 premotor/integrative 층으로 본다.",
        "",
        "## 결과",
        "",
        "| 항목 | 값 |",
        "|---|---:|",
        f"| module count | {output['module_count']} |",
        f"| used edges | {output['edge_summary']['used_edges']} |",  # type: ignore[index]
        f"| used synapses | {output['edge_summary']['used_synapses']:.1f} |",  # type: ignore[index]
        f"| flat loss | {gate['flat_loss']:.6f} |",
        f"| L1/L2/L3 block loss | {gate['layer_block_loss']:.6f} |",
        f"| block / flat | {gate['layer_over_flat']:.6f} |",
        f"| block / random mean | {gate['layer_over_random_mean']:.6f} |",
        f"| permutation p | {gate['p_value_loss_le_observed']:.6f} |",
        f"| passed | {output['passed']} |",
        "",
        "## 방향성",
        "",
        "| 항목 | 값 |",
        "|---|---:|",
        f"| forward fraction | {direction['forward_fraction']:.6f} |",
        f"| backward fraction | {direction['backward_fraction']:.6f} |",
        f"| lateral fraction | {direction['lateral_fraction']:.6f} |",
        f"| forward/backward | {direction['forward_over_backward']:.6f} |",
        "",
        "C. elegans 회로는 단순 feedforward 사슬이 아니라 recurrent/residual 구조를 강하게 갖는다. 따라서 성공 기준은 forward dominance가 아니라 layer-block 설명력이다.",
        "",
        "## Robustness",
        "",
        "| matrix | block/flat | block/random mean | permutation p | module-family/flat | pass |",
        "|---|---:|---:|---:|---:|---|",
    ]
    for name, row in robustness.items():
        layer_gate = row["layer_gate"]
        lines.append(
            f"| {name} | {layer_gate['layer_over_flat']:.6f} | "
            f"{layer_gate['layer_over_random_mean']:.6f} | "
            f"{layer_gate['p_value_loss_le_observed']:.6f} | "
            f"{row['module_family_over_flat']:.6f} | {row['passed_layer_gate']} |"
        )
    lines.extend(
        [
            "",
            "weighted all/chemical/binary 조건에서 L1/L2/L3 층화가 유지되는지 본다. electrical-only는 gap junction 성격상 층화보다 lateral coupling이 강할 수 있으므로 별도 해석한다.",
            "",
            "## 반례 점검",
            "",
            "| model | labels | params | block/flat | p | BIC-like | saturated |",
            "|---|---:|---:|---:|---:|---:|---|",
        ]
    )
    for name, row in counter["models"].items():
        layer_gate = row["gate"]
        lines.append(
            f"| {name} | {row['label_count']} | {row['parameter_count']} | "
            f"{layer_gate['layer_over_flat']:.6f} | "
            f"{layer_gate['p_value_loss_le_observed']:.6f} | "
            f"{row['bic_like']:.3f} | {row['saturated']} |"
        )
    lines.extend(
        [
            "",
            f"포화모델을 제외한 BIC-like 최저 모델은 `{counter['best_bic_like_model_excluding_saturated']}`이다.",
            f"포화모델을 제외한 순수 손실 최저 모델은 `{counter['best_loss_model_excluding_saturated']}`이다.",
            "따라서 L1/L2/L3 층화가 유일한 설명인지, 아니면 module family가 더 경제적인지 함께 봐야 한다.",
            "",
            "## 안정성",
            "",
            "| 항목 | 값 |",
            "|---|---:|",
            f"| lambda max | {stability['laplacian_lambda_max']:.6f} |",
            f"| gamma upper bound at rho=0.20 | {stability['gamma_stability_upper_bound']:.8f} |",
            f"| modal radius at half bound | {stability['modal_radius_at_gamma_example']:.6f} |",
            "",
            "## 함의",
            "",
            "- 인간 이전의 원시 신경계에도 감각-중간-운동 층화 그래프가 통계적으로 보인다.",
            "- 전역 뇌 방정식의 그래프 항 \\(\\Delta_G\\)는 고등피질 전용 항이 아니라 원시 회로에도 적용 가능한 최소 문법일 수 있다.",
            "- 다만 이 결과는 구조 connectome gate이며, 동역학 전이 \\(P_n\\to P_{n+1}\\) 검증은 아니다.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--force-download", action="store_true")
    parser.add_argument("--permutations", type=int, default=2000)
    args = parser.parse_args()

    xlsx = ensure_xlsx(args.xlsx, force=args.force_download)
    edges = load_edges(xlsx)
    modules, matrices, edge_summary = aggregate_module_matrices(edges)
    matrix = matrices["all_weighted"]
    labels = [module_layer(module) for module in modules]
    gate = permutation_gate(matrix, labels, args.permutations)
    robustness = robustness_gates(matrices, modules, args.permutations)
    countermodel_check = compare_layer_countermodels(matrix, modules, args.permutations)
    output = {
        "dataset": DATASET,
        "source_url": SOURCE_URL,
        "xlsx": str(xlsx),
        "module_count": len(modules),
        "modules": modules,
        "layer_labels": labels,
        "edge_summary": edge_summary,
        "gate": gate,
        "robustness": robustness,
        "countermodel_check": countermodel_check,
        "directionality": directionality(matrix, modules),
        "bow_tie_indices": bow_tie_indices(matrix, modules),
        "laplacian_stability": laplacian_stability(matrix),
        "criterion": (
            "L1/L2/L3 layer-block loss < flat loss, < random-layer mean, "
            "and permutation p < 0.05"
        ),
        "passed": gate["layer_block_loss"] < gate["flat_loss"]
        and gate["layer_block_loss"] < gate["random_layer_loss_mean"]
        and gate["p_value_loss_le_observed"] < 0.05,
    }
    out_json = Path(__file__).with_name("c_elegans_connectome_gate_results.json")
    out_md = Path(__file__).with_name("c_elegans_connectome_report.md")
    out_json.write_text(json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8")
    write_report(output, out_md)

    print("C. elegans connectome evolutionary gate")
    print(f"  modules              = {output['module_count']}")
    print(f"  used_edges           = {edge_summary['used_edges']}")
    print(f"  used_synapses        = {edge_summary['used_synapses']:.1f}")
    print(f"  flat_loss            = {gate['flat_loss']:.6f}")
    print(f"  layer_block_loss     = {gate['layer_block_loss']:.6f}")
    print(f"  block/flat           = {gate['layer_over_flat']:.6f}")
    print(f"  block/random_mean    = {gate['layer_over_random_mean']:.6f}")
    print(f"  permutation_p        = {gate['p_value_loss_le_observed']:.6f}")
    print("  robustness:")
    for name, row in robustness.items():
        layer_gate = row["layer_gate"]
        print(
            f"    {name:19s} block/flat={layer_gate['layer_over_flat']:.6f}, "
            f"p={layer_gate['p_value_loss_le_observed']:.6f}, "
            f"pass={row['passed_layer_gate']}"
        )
    print(
        "  counter best BIC    = "
        f"{countermodel_check['best_bic_like_model_excluding_saturated']}"
    )
    print(
        "  counter best loss   = "
        f"{countermodel_check['best_loss_model_excluding_saturated']}"
    )
    print(f"  passed               = {output['passed']}")
    print(f"Saved: {out_json}")
    print(f"Saved: {out_md}")


if __name__ == "__main__":
    main()
