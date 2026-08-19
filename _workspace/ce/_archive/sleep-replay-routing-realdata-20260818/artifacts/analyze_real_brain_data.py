from __future__ import annotations

import csv
import hashlib
import json
import subprocess
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.io import loadmat
from scipy.stats import spearmanr


HERE = Path(__file__).resolve().parent
REALDATA = HERE / "realdata"


def load_frame(path: Path) -> dict[str, dict[int, object]]:
    payload = np.load(path, allow_pickle=True).item()
    if set(payload) != {"data"} or not isinstance(payload["data"], dict):
        raise ValueError(f"unexpected GroupData schema: {path}")
    return payload["data"]


def column(data: dict[str, dict[int, object]], name: str) -> np.ndarray:
    values = data[name]
    return np.asarray([values[i] for i in sorted(values)])


def bootstrap_prob(sample1: np.ndarray, sample2: np.ndarray) -> tuple[float, float]:
    """Byte-for-byte translation of the repository's get_bootstrap_prob."""
    low = min(sample1.min(), sample2.min())
    high = max(sample1.max(), sample2.max())
    axis = np.linspace(low, high, num=100)
    edge_shift = (axis[2] - axis[1]) / 2
    edges = np.append(axis - edge_shift, high + edge_shift)
    p1 = np.histogram(sample1, bins=edges)[0] / sample1.size
    p2 = np.histogram(sample2, bins=edges)[0] / sample2.size
    joint = p1[:, None] * p2[None, :]
    joint /= joint.sum()
    probability = np.triu(joint).sum()
    probability = 1 - probability if probability >= 0.5 else probability
    statistic = abs(sample1.mean() - sample2.mean()) / np.sqrt(
        (sample1.std() ** 2 + sample2.std() ** 2) / 2
    )
    return float(statistic), float(probability)


def e15_replay() -> dict[str, object]:
    root = REALDATA / "e15_repo"
    processed = root / "ProcessedData"
    raw = load_frame(processed / "continuous_replay_number_1h_blocks.npy")
    boot = load_frame(processed / "continuous_replay_number_1h_blocks_bootstrap.npy")

    raw_grp = column(raw, "grp")
    raw_zt = column(raw, "zt")
    raw_value = column(raw, "is_cont").astype(float)
    boot_grp = column(boot, "grp")
    boot_zt = column(boot, "zt")
    boot_value = column(boot, "is_cont").astype(float)

    epochs: dict[str, object] = {}
    for epoch in ("0-1", "4-5", "5-6"):
        values = {}
        samples = {}
        for group in ("NSD", "SD"):
            raw_mask = (raw_grp == group) & (raw_zt == epoch)
            boot_mask = (boot_grp == group) & (boot_zt == epoch)
            values[group] = {
                "n_sessions": int(raw_mask.sum()),
                "mean_events": float(raw_value[raw_mask].mean()),
                "sd_events": float(raw_value[raw_mask].std(ddof=1)),
            }
            samples[group] = boot_value[boot_mask]
        statistic, probability = bootstrap_prob(samples["NSD"], samples["SD"])
        epochs[epoch] = {
            "session_values": values,
            "official_bootstrap_statistic": statistic,
            "official_bootstrap_probability": probability,
            "direction_sd_minus_nsd": values["SD"]["mean_events"]
            - values["NSD"]["mean_events"],
        }

    commit = subprocess.check_output(
        ["git", "-C", str(root), "rev-parse", "HEAD"], text=True
    ).strip()
    return {
        "repository_commit": commit,
        "processed_rows": int(raw_value.size),
        "released_session_labels": int(np.unique(column(raw, "session")).size),
        "independence_status": "NOT_ESTABLISHED_FROM_PROCESSED_TABLE",
        "epochs": epochs,
        "paper_notebook_checks": {
            "0-1_probability": 0.01611,
            "5-6_probability": 0.001337,
            "source": "sd_figure4_bs.ipynb execution_count 12",
        },
        "same_window_branching_replay": {
            "status": "UNTESTABLE",
            "reason": (
                "The release contains processed session/hour replay and firing-rate summaries, "
                "not linked unit spike timestamps, SWR events, and replay scores in the same windows."
            ),
        },
    }


def matlab_cluster(path: Path) -> np.ndarray:
    payload = loadmat(path)
    selected = payload["P_cluster_idx"].ravel().astype(int)
    if selected.size != 1:
        raise ValueError(f"expected one selected cluster in {path}")
    # MATLAB indices in both the cell array and coordinate list are one-based.
    return payload["clusters_idx"][selected[0] - 1, 0].astype(int) - 1


def e19_sleep_geometry() -> dict[str, object]:
    root = REALDATA / "e19_data"
    retrieval = loadmat(root / "closeeye_corr_conds_final.mat")["closeeye_corr_suball"]
    encoding = loadmat(root / "encode_corr_conds_final.mat")["encode_corr_suball"]
    sleep = loadmat(root / "suball_neural_sleep_param.mat")[
        "suball_neural_sleep_rem_cluster"
    ]

    subjects = [i for i in range(35) if i != 4]

    def adjusted(first: int, second: int) -> np.ndarray:
        r = np.stack(
            [retrieval[i, first] - retrieval[i, second] for i in subjects]
        )
        e = np.stack([encoding[i, first] - encoding[i, second] for i in subjects])
        baseline = e.mean(axis=(1, 2))
        return r - baseline[:, None, None]

    # MATLAB conditions [5 6] for item and [6 7] for category.
    item = adjusted(4, 5)
    category = adjusted(5, 6)
    ratio = sleep[:, 1] / sleep[:, 2]

    outcomes = {}
    for name, cube, cluster_file in (
        ("item", item, "cluster_permute_test_RRS_item_rem2sws.mat"),
        ("category", category, "cluster_permute_test_RRS_cate_rem2sws.mat"),
    ):
        coords = matlab_cluster(root / cluster_file)
        y = cube[:, coords[:, 0], coords[:, 1]].mean(axis=1)
        rho, pvalue = spearmanr(ratio, y)
        slope, intercept = np.polyfit(ratio, y, 1)
        outcomes[name] = {
            "n_participants": int(y.size),
            "cluster_pixels": int(coords.shape[0]),
            "spearman_rho": float(rho),
            "spearman_p_two_sided": float(pvalue),
            "ols_slope_descriptive": float(slope),
            "ols_intercept_descriptive": float(intercept),
        }
    return {
        "participants": 34,
        "official_exclusion": "MATLAB participant 5: disconnected sleep EEG",
        "outcomes": outcomes,
        "scope": (
            "Participant-level reproduction of the official Figure 4 cluster association. "
            "It is not a direct measurement of dream content or generative recombination."
        ),
    }


def sheet_numeric(path: Path, sheet: str) -> tuple[list[str], np.ndarray]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = list(workbook[sheet].iter_rows(values_only=True))
    header = [str(value) for value in rows[0]]
    return header, np.asarray(rows[1:], dtype=float)


def e13_criticality() -> dict[str, object]:
    root = REALDATA / "e13_repo"
    source = root / "Source Data" / "Source Data Fig 2.xlsx"
    header_d, fig2d = sheet_numeric(source, "fig 2d")
    header_e, fig2e = sheet_numeric(source, "fig 2e")

    def describe(header: list[str], data: np.ndarray, a: str, b: str) -> dict[str, float]:
        av = data[:, header.index(a)]
        bv = data[:, header.index(b)]
        return {
            f"mean_{a}": float(av.mean()),
            f"mean_{b}": float(bv.mean()),
            f"mean_difference_{a}_minus_{b}": float((av - bv).mean()),
        }

    workbook = load_workbook(source, read_only=True, data_only=True)
    stats_rows = list(workbook["fig 2d stats"].iter_rows(values_only=True))
    dcc_vs_shuffle = next(
        row for row in stats_rows if row[1] == "DCC" and row[2] == "shuffle"
    )
    commit = subprocess.check_output(
        ["git", "-C", str(root), "rev-parse", "HEAD"], text=True
    ).strip()
    return {
        "repository_commit": commit,
        "fig2d_prediction": describe(header_d, fig2d, "DCC", "Shuffle"),
        "fig2d_source_table_tukey": {
            "mean_difference": float(dcc_vs_shuffle[3]),
            "p_adjusted": float(dcc_vs_shuffle[4]),
            "lower": float(dcc_vs_shuffle[5]),
            "upper": float(dcc_vs_shuffle[6]),
            "reject": bool(dcc_vs_shuffle[7]),
        },
        "fig2e_increment": describe(header_e, fig2e, "DCC+Base", "Base"),
        "scope": (
            "Deterministic reproduction from published source tables. Raw broadband exceeds "
            "10 TB and remains author-request only, so no new event-level branching estimate is made."
        ),
    }


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_manifest() -> int:
    files = []
    for path in sorted(REALDATA.rglob("*")):
        if not path.is_file() or ".git" in path.parts:
            continue
        files.append(
            {
                "relative_path": path.relative_to(HERE).as_posix(),
                "bytes": path.stat().st_size,
                "sha256": sha256(path),
            }
        )
    with (HERE / "realdata-manifest.csv").open("w", newline="", encoding="utf-8") as out:
        writer = csv.DictWriter(out, fieldnames=("relative_path", "bytes", "sha256"))
        writer.writeheader()
        writer.writerows(files)
    return len(files)


def main() -> None:
    results = {
        "e15": e15_replay(),
        "e19": e19_sleep_geometry(),
        "e13": e13_criticality(),
        "e02": {
            "status": "ACCESS_BLOCKED",
            "reason": (
                "Dryad metadata are public, but file streams returned HTTP 403 and the v2 file "
                "download endpoint required an OAuth bearer token. No 56-byte error response was "
                "treated as neural data."
            ),
            "predeclared_sessions": ["m1_ses1", "m1_ses2", "m1_ses24", "m1_ses25"],
        },
    }
    results["manifest_file_count"] = write_manifest()
    output = HERE / "realdata-results.json"
    output.write_text(json.dumps(results, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    print(json.dumps(results, indent=2, ensure_ascii=True))


if __name__ == "__main__":
    main()
